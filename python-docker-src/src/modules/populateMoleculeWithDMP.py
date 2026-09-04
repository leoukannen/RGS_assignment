from __future__ import annotations

import os
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pymongo import MongoClient

from modules.defineMoleculeDetailsTable import MOLECULE_DETAILS_COLLECTION


DMP_WORKBOOK_PATH = Path(
	os.getenv("DMP_WORKBOOK_PATH", "/app/data/dmp/maximum-prices.xlsx")
)
DMP_SOURCE_URL = (
	"https://www.dmp.no/offentlig-finansiering/pris-pa-legemidler/maksimalpris"
)


def populate_molecule_with_dmp(
	client: MongoClient[dict[str, Any]],
	workbook_path: Path = DMP_WORKBOOK_PATH,
) -> int:
	"""Fill null moleculeDetails fields with matching DMP maximum-price data."""
	price_rows = _read_price_rows(workbook_path)
	collection = client.get_default_database()[MOLECULE_DETAILS_COLLECTION]
	updated = 0

	for document in collection.find({}):
		row = _find_matching_row(document, price_rows)
		if row is None:
			continue
		updates = _null_field_updates(document, row)
		updates.update(_source_updates(document))
		if updates:
			collection.update_one({"_id": document["_id"]}, {"$set": updates})
			updated += 1

	print(f"Updated {updated} molecule detail(s) with DMP data")
	return updated


def _read_price_rows(path: Path) -> dict[str, list[dict[str, Any]]]:
	if not path.is_file():
		raise FileNotFoundError(f"DMP workbook not found: {path}")

	rows: dict[str, list[dict[str, Any]]] = {}
	workbook = load_workbook(path, read_only=True, data_only=True)
	try:
		worksheet = workbook.active
		row_iterator = worksheet.iter_rows(values_only=True)
		headers: list[str] = []
		for values in row_iterator:
			candidate_headers = [
				str(value).strip() if value is not None else "" for value in values
			]
			if "Varenummer" in candidate_headers:
				headers = candidate_headers
				break
		if not headers:
			raise ValueError("DMP workbook has no Varenummer header")
		while True:
			values = next(row_iterator, None)
			if values is None:
				break
			if not any(value is not None for value in values):
				continue
			row = dict(zip(headers, values))
			item_number = _clean(row.get("Varenummer"))
			if item_number:
				rows.setdefault(item_number, []).append(row)
	finally:
		workbook.close()
	return rows


def _find_matching_row(
	document: dict[str, Any],
	price_rows: dict[str, list[dict[str, Any]]],
) -> dict[str, Any] | None:
	item_number = _clean(document.get("itemNumber"))
	if item_number in price_rows:
		matching_rows = price_rows[item_number]
		for row in matching_rows:
			if _same_atc(document, row):
				return row
		return matching_rows[0]
	return None


def _same_atc(document: dict[str, Any], row: dict[str, Any]) -> bool:
	document_atc = _clean(document.get("atcCode"))
	row_atc = _clean(row.get("ATC-kode (pakning)"))
	return bool(document_atc and row_atc and document_atc == row_atc)


def _null_field_updates(
	document: dict[str, Any], row: dict[str, Any]
) -> dict[str, Any]:
	values = {
		"moleculeVariant": _clean(row.get("Virkestoff")),
		"productName": _clean(row.get("Handelsnavn")),
		"strength": _clean(row.get("Styrke")),
		"packSize": _pack_size(row),
		"supplier": _clean(row.get("Innehaver")),
		"maxPrice": _clean(row.get("Maks AUP Gyldig")),
		"currency": "NOK",
		"status": _clean(row.get("Markedsføringsstatus")),
		"publicationDate": _excel_date(row.get("AIP Gyldig Fradato")),
	}
	return {
		field: value
		for field, value in values.items()
		if _is_null(document.get(field)) and value is not None
	}


def _pack_size(row: dict[str, Any]) -> str | None:
	amount = _clean(row.get("Mengde per beholder"))
	unit = _clean(row.get("Måle-enhet"))
	if amount and unit:
		return f"{amount} {unit}"
	return amount or _clean(row.get("Antall beholdere"))


def _excel_date(value: Any) -> str | None:
	if isinstance(value, datetime):
		return value.date().isoformat()
	if isinstance(value, (int, float)):
		return (datetime(1899, 12, 30) + timedelta(days=value)).date().isoformat()
	return _clean(value)


def _clean(value: Any) -> str | None:
	if value is None:
		return None
	value = str(value).strip()
	return value or None


def _is_null(value: Any) -> bool:
	return value is None or value == ""


def _source_updates(document: dict[str, Any]) -> dict[str, list[str]]:
	return {
		"sourceDocument": _append_unique(
			document.get("sourceDocument"), "DMP maksimalpriser"
		),
		"sourceUrl": _append_unique(document.get("sourceUrl"), DMP_SOURCE_URL),
	}


def _append_unique(value: Any, addition: str) -> list[str]:
	if isinstance(value, list):
		values = [str(item) for item in value if item is not None]
	elif value is None or value == "":
		values = []
	else:
		values = [str(value)]
	if addition not in values:
		values.append(addition)
	return values